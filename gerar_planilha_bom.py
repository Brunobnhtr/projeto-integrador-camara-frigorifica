#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera a planilha de compras do Projeto Integrador a partir da BOM em markdown.

A lista de compras parte da BOM em Markdown e recebe uma conferencia cruzada
com os dados renderizados pelo React nas abas de inventario, fiacao e
divergencias.

Este script le as tabelas daquele documento, monta a lista principal e registra
as decisoes que ainda precisam ser reconciliadas entre BOM e React.

    pip install openpyxl
    python gerar_planilha_bom.py

Substitui os antigos gerar_excel_materiais.py e gerar_excel_completo.py, que
tinham a lista escrita na mão e ainda continham a fonte ATX e o disjuntor
removidos na refatoração para 24 V.
"""

from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Falta a biblioteca openpyxl. Instale com:  pip install openpyxl")

RAIZ = Path(__file__).resolve().parent
BOM_MD = RAIZ / "camada_0_fundamentos" / "03_lista_materiais.md"
SAIDA = RAIZ / "BOM_Projeto_Integrador.xlsx"

# ── paleta (mesma identidade dos desenhos e dos vídeos) ────────────────────
AZUL_ESCURO = "1D3557"
AMBAR = "F5A524"
CINZA_CLARO = "EFF2F5"
CINZA_LINHA = "D6DCE4"
VERDE = "2B8A3E"

FINA = Side(style="thin", color=CINZA_LINHA)
BORDA = Border(left=FINA, right=FINA, top=FINA, bottom=FINA)

# Seções da BOM que descrevem o que SAIU do projeto — nunca viram compra.
SECAO_REMOVIDOS = re.compile(r"n[ãa]o existem mais|removid", re.IGNORECASE)


def limpar(texto: str) -> str:
    """Tira a formatação markdown, deixando o texto puro para a planilha."""
    texto = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", texto)  # [rótulo](link) -> rótulo
    texto = texto.replace("**", "").replace("`", "")
    texto = texto.replace("~~", "")
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


# ── links de busca ─────────────────────────────────────────────────────────
#
# Por que links de BUSCA e não links de anúncio: anúncio do Mercado Livre sai
# do ar, muda de vendedor e muda de preço. Um link de busca continua válido
# para sempre e sempre mostra o que existe HOJE. A coluna "Link escolhido"
# fica ao lado, para você colar o anúncio depois de decidir.

# Palavras que não ajudam a achar o produto — poluem a busca.
RUIDO = {
    "de", "da", "do", "das", "dos", "para", "com", "sem", "e", "ou", "a", "o",
    "as", "os", "em", "no", "na", "por", "cada", "tipo", "opcional",
}
ACENTOS = str.maketrans("áàâãäéèêëíìîïóòôõöúùûüçÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ",
                        "aaaaaeeeeiiiiooooouuuucAAAAAEEEEIIIIOOOOOUUUUC")


def termo_de_busca(item: str) -> str:
    """Transforma o nome do item num termo de busca utilizável numa loja.

    Devolve "" para linhas que não são itens de compra — as que começam com
    "→" detalham o item da linha anterior (ex.: "→ T2 (poste P2)") e gerariam
    uma busca inútil por "t2".
    """
    if item.lstrip().startswith(("→", "->")):
        return ""
    texto = item.translate(ACENTOS).lower()
    texto = re.sub(r"\(.*?\)", " ", texto)          # tira parênteses explicativos

    def palavras_de(t: str) -> list[str]:
        t = re.sub(r"[^a-z0-9,\s]", " ", t)
        return [p for p in t.split() if p not in RUIDO]

    # Normalmente o que interessa está ANTES do travessão ("Fonte — 24 V ...").
    # Mas em itens como "KM1 — Relé de interface 24 Vcc" o rótulo do projeto
    # vem antes e o nome do produto vem depois. Se o começo render menos de
    # 2 palavras úteis, ele é rótulo: usa-se o nome completo.
    palavras = palavras_de(texto.split("—")[0].split("·")[0])
    if len(palavras) < 2:
        palavras = palavras_de(texto.replace("—", " ").replace("·", " "))
        # "KM1", "T2", "F1" são identificadores do projeto, não do produto.
        # Numa busca em loja eles só atrapalham.
        while palavras and re.fullmatch(r"[a-z]{1,3}\d{1,2}", palavras[0]):
            palavras.pop(0)

    return " ".join(palavras[:7])                    # 7 palavras já é específico


def url_mercadolivre(termo: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", termo.lower()).strip("-")
    return f"https://lista.mercadolivre.com.br/{slug}" if slug else ""


def url_amazon(termo: str) -> str:
    from urllib.parse import quote_plus
    return f"https://www.amazon.com.br/s?k={quote_plus(termo)}" if termo else ""


# ── AliExpress ─────────────────────────────────────────────────────────────
#
# Duas diferenças em relação às lojas nacionais:
#
# 1. A busca precisa ser em INGLÊS. Termo em português não acha quase nada.
# 2. Nem todo item DEVE ser importado. Componentes que fazem parte da
#    segurança elétrica (disjuntor, relés, fonte) devem ser nacionais e
#    certificados — o edital exige "conformidade com normas de segurança
#    elétrica", e produto sem INMETRO não sustenta essa afirmação.
#
# Por isso a lista abaixo é CURADA: só aparece link de AliExpress no item
# que faz sentido importar. Ausência de link aqui é recomendação, não falha.
BUSCA_ALIEXPRESS = {
    # Reles: so entram com MARCA e MODELO. O criterio nao e "importado e
    # ruim" - e que o datasheet precisa declarar a corrente em DC, e
    # anuncio generico ("24V 10A Relay Module") nao declara.
    "km1":                       "JQX-13F LY2N 24VDC relay 10A 8 pin PTF08A socket",
    "kit duplo de refrigera":     "dual TEC1-12706 peltier cooler kit 12V",
    "peltier tec1-12706":        "TEC1-12706 peltier module",
    "ventoinha de reposi":       "80mm fan 12V 3 pin PWM tachometer",
    "aquecedor ptc":             "24V PTC heater element with fan",
    "sensor ina219":             "INA219 current sensor module",
    "uln2803":                   "ULN2803A DIP18 darlington",
    "soquete dip":               "DIP18 IC socket",
    "kit 2":                     "LM2596 buck converter voltmeter display",
    "kit lm2596":                "LM2596 buck converter voltmeter display",
    "bts7960":                   "BTS7960 43A motor driver module",
    "esp32-wroom":               "ESP32 WROOM 32U",
    "arduino mega":              "Arduino Mega 2560 R3",
    "nextion":                   "Nextion NX4024T032 3.2 inch HMI",
    "ds18b20":                   "DS18B20 waterproof temperature sensor",
    "am2315c":                   "AM2315C temperature humidity sensor",
    "rtc ds3231":                "DS3231 RTC module",
    "micro sd":                  "micro SD card module SPI arduino",
    "cartão micro sd":           "micro SD card 16GB class 10",
    "dissipador + cooler 80":    "80mm heatsink fan CPU cooler 3pin",
    "dissipador de alumínio":    "aluminum heatsink small",
    "cooler 60 mm":              "60mm fan 24V DC",
    "cooler 40":                 "40mm fan 12V DC",
    "fan 60":                    "60mm fan 12V DC",
    "fan 40":                    "40mm fan 12V DC",
    "pasta térmica":             "thermal paste syringe",
    "fita térmica":              "thermal adhesive tape double sided",
    "placa ilhada":              "perfboard prototype PCB",
    "borne kf350":               "KF350 3.5mm PCB screw terminal block",
    "espaguete termorretrátil":  "heat shrink tubing kit",
    "termorretrátil":            "heat shrink tubing kit",
    "sinaleiros led 22":         "22mm LED pilot indicator lamp 24V",
    "botão de emergência":       "22mm emergency stop mushroom button",
    "botão start":               "22mm push button switch green",
    "botão stop":                "22mm push button switch black",
    "blocos de contato":         "22mm push button contact block NO NC",
    "voltímetro + amperímetro":  "DC voltmeter ammeter 100V 10A digital",
    "pigtail":                   "IPEX u.FL to SMA pigtail cable",
    "antena 2,4":                "2.4GHz SMA antenna 3dBi",
    "conector sma":              "SMA bulkhead panel connector female",
    "conector circular gx16":    "GX16 8 pin aviation connector",
    "resistor":                  "1/4W resistor kit assorted",
    "capacitor cerâmico":        "100nF ceramic capacitor",
    "diodo 1n4007":              "1N4007 diode",
    "led 3 mm":                  "3mm warm white LED",
    "led 5 mm":                  "5mm diffused LED",
    "micro-chave":               "SPDT micro switch small",
}


# Acessórios que CITAM o componente principal no nome mas são outra coisa.
# Sem isto, "Suporte DIN para Arduino Mega" cai na busca de "Arduino Mega".
ACESSORIO = re.compile(r"^\s*(suporte|base|flange|tampa|grade|cinta)\b", re.I)


def busca_aliexpress(item: str) -> str:
    """Devolve o termo em inglês se o item valer a pena importar."""
    if ACESSORIO.match(item.translate(ACENTOS)):
        return ""
    chave = item.translate(ACENTOS).lower()
    alvo = {k.translate(ACENTOS): v for k, v in BUSCA_ALIEXPRESS.items()}
    for fragmento, termo_en in alvo.items():
        if fragmento in chave:
            return termo_en
    return ""


def url_aliexpress(termo_en: str) -> str:
    if not termo_en:
        return ""
    slug = re.sub(r"[^a-z0-9]+", "-", termo_en.lower()).strip("-")
    return f"https://pt.aliexpress.com/w/wholesale-{slug}.html"


def linha_de_tabela(linha: str) -> list[str] | None:
    linha = linha.strip()
    if not linha.startswith("|") or not linha.endswith("|"):
        return None
    return [c.strip() for c in linha[1:-1].split("|")]


def eh_separador(celulas: list[str]) -> bool:
    return all(re.fullmatch(r":?-{2,}:?", c) for c in celulas if c)


def ler_bom(caminho: Path) -> list[dict]:
    """Extrai as tabelas de materiais, guardando a seção de cada uma."""
    if not caminho.exists():
        sys.exit(f"BOM não encontrada: {caminho}")

    itens: list[dict] = []
    secao = "Geral"
    cabecalho: list[str] | None = None

    for linha in caminho.read_text(encoding="utf-8").splitlines():
        titulo = re.match(r"^#{1,3}\s+(.*)", linha)
        if titulo:
            secao = limpar(titulo.group(1))
            cabecalho = None
            continue

        celulas = linha_de_tabela(linha)
        if celulas is None:
            cabecalho = None
            continue
        if eh_separador(celulas):
            continue

        # primeira linha de uma tabela = cabeçalho
        if cabecalho is None:
            cabecalho = [limpar(c) for c in celulas]
            continue

        # só interessam as tabelas que descrevem materiais
        if not cabecalho or "Item" not in cabecalho[0]:
            continue

        # ⚠️ A tabela "Itens que NÃO existem mais" também tem "Item" no
        #    cabeçalho ("Item removido"). Sem este filtro, componentes já
        #    eliminados do projeto — XL4016, Zener, cooler do T1 — voltam
        #    para a lista de compras.
        if SECAO_REMOVIDOS.search(secao):
            continue

        brutos = dict(zip(cabecalho, celulas))
        valores = dict(zip(cabecalho, [limpar(c) for c in celulas]))
        item = valores.get(cabecalho[0], "")

        # O tachado (~~item~~) marca o que saiu do projeto. A checagem é feita
        # na célula BRUTA porque limpar() remove os "~~" antes de chegar aqui.
        if not item or "~~" in brutos.get(cabecalho[0], ""):
            continue

        # a 4ª coluna muda de nome conforme a tabela (Busca / Ajuste / Aplicação)
        extras = [
            valores.get(k, "")
            for k in cabecalho[3:]
            if k.lower() not in ("link", "link de compra")
        ]

        # "Busca sugerida" da BOM tem prioridade sobre o nome do item, porque
        # foi escrita pensando em como o produto é anunciado, não em como ele
        # é chamado no projeto.
        termo = valores.get("Busca sugerida", "") or termo_de_busca(item)

        itens.append(
            {
                "secao": secao,
                "item": item,
                "qtd": valores.get("Qtd", ""),
                "espec": valores.get("Especificação", ""),
                "obs": " · ".join(x for x in extras if x),
                "link_manual": limpar(brutos.get("Link", "")),
                "busca_ml": url_mercadolivre(termo),
                "busca_amz": url_amazon(termo),
                "termo": termo,
                "termo_en": busca_aliexpress(item),
                "busca_ali": url_aliexpress(busca_aliexpress(item)),
            }
        )

    return itens


# Inventario conferido nos dados que o React realmente renderiza. Esta aba nao
# substitui a BOM: ela torna visiveis as decisoes do modelo que ainda nao estao
# representadas de forma confiavel nas tabelas Markdown.
REACT_INVENTARIO = [
    ("Painel", "Caixa do painel", 1, "500 x 500 x 200 mm", "painel_interativo/src/data/painel_completo.js"),
    ("Painel", "BD-POT", 1, "24 V de potencia, 1 entrada + 4 saidas", "painel_interativo/src/data/painel.js"),
    ("Painel", "BD-AUX", 1, "12 V auxiliar, 1 entrada + 4 saidas", "painel_interativo/src/data/painel.js"),
    ("Painel", "BD-24V", 1, "24 V permanente, 1 entrada + 6 saidas", "painel_interativo/src/data/painel_completo.js"),
    ("Painel", "BD-5V", 1, "5 V, 1 entrada + 8 saidas", "painel_interativo/src/data/painel_completo.js"),
    ("Painel", "BD-0V", 1, "barra star ground, minimo 20 pontos", "painel_interativo/src/data/painel_completo.js"),
    ("Painel", "BTS7960", 2, "1 para Peltier; 1 para PTC", "painel_interativo/src/data/painel_completo.js"),
    ("Painel", "Rele 8 pinos + base DIN", 2, "KM1 (1 em uso + 1 reserva), bobina 24 Vcc, contato 10 A em CC", "painel_interativo/src/data/reles_fisico.js"),
    ("Painel", "Modulo rele 1 canal 5 V", 3, "KA1 NO, KA2 NC e KA3 NO, optoacoplados; caixa DIN 6M", "painel_interativo/src/data/reles_fisico.js"),
    ("Painel", "Arduino Mega 2560", 1, "Controlador principal", "painel_interativo/src/data/painel_completo.js"),
    ("Painel", "Placa PI-1", 1, "Placa ilhada com ULN2803A, filtros e divisor D25", "painel_interativo/src/data/pi1_fisico.js"),
    ("Painel", "Placa PI-2", 1, "Placa ilhada com CD74HC4067, shunts e INA219", "painel_interativo/src/data/pi2_fisico.js"),
    ("Painel", "DNLCB30 + ESP32-WROOM-32U", 1, "Supervisao Wi-Fi/MQTT; entrada 24 V", "painel_interativo/src/data/painel_completo.js"),
    ("Painel", "ES3C28P ESP32-S3", 1, "IHM 2,8 pol, touch, microSD; alimentacao pelo Type-C", "painel_interativo/src/data/pinagens.js"),
    ("Painel", "Conversor de nivel UART", 1, "2 canais, 5 V <-> 3,3 V", "painel_interativo/src/data/pinagens.js"),
    ("Painel", "Sinaleiro LED 22 mm 24 V", 4, "RUN verde, COOL azul, HEAT amarelo, FAULT vermelho", "painel_interativo/src/data/painel.js"),
    ("Painel", "Botoeiras", 3, "Emergencia (cogumelo), LIGAR verde e STOP preto", "painel_interativo/src/data/painel.js"),
    ("Camara", "Peltier TEC1-12706", 2, "Ligadas em serie, carga de 24 V / aproximadamente 6 A", "painel_interativo/src/data/camara.js"),
    ("Camara", "PTC 24 V", 1, "Aquecedor; comandado pelo BTS #2", "painel_interativo/src/data/camara.js"),
    ("Camara", "Ventoinhas internas 12 V", 5, "4 de circulacao + 1 do PTC, em paralelo", "painel_interativo/src/data/camara_ligacoes.js"),
    ("Fora da camara", "Cooler do radiador 3 fios", 2, "12 V, com RPM individual; lado quente", "painel_interativo/src/data/camara_ligacoes.js"),
    ("Camara", "AM2315C", 1, "Temperatura e umidade, dentro da camara", "painel_interativo/src/data/camara.js"),
    ("Fora da camara", "DS18B20", 1, "Temperatura do dissipador quente, 3 fios", "painel_interativo/src/data/camara_ligacoes.js"),
    ("Camara", "Posicao de ensaio / DUT", 2, "LED + resistor; retornos individuais RET-1 e RET-2", "painel_interativo/src/data/camara.js"),
    ("Camara", "Porta-fusivel das posicoes", 1, "2 vias, fusivel 100 mA por DUT", "painel_interativo/src/data/camara_ligacoes.js"),
    ("Camara", "Modulo CD74HC4067", 1, "Multiplexador de 16 canais", "painel_interativo/src/data/pi2_fisico.js"),
    ("Camara", "Modulo INA219", 1, "Medicao de referencia da posicao 1", "painel_interativo/src/data/pi2_fisico.js"),
]

REACT_FIACAO = [
    ("127 V CA", "preto / azul / verde-amarelo", "1,5 mm2", "Tomada -> disjuntor -> fonte; PE aterra a carcaça", "painel_interativo/src/data/maquete.js"),
    ("24 V potencia", "vermelho", "1,5 mm2", "P1/painel -> KM1 -> BD-POT -> BTS", "painel_interativo/src/data/fiacao.js"),
    ("0 V comum", "preto ou azul escuro", "1,5 mm2", "Retorno geral e star ground BD-0V", "painel_interativo/src/data/fiacao.js"),
    ("24 V linha R1", "vermelho rigido encapado", "1,00 mm2", "Ramal das Peltier, 6,0 A", "painel_interativo/src/data/maquete.js"),
    ("24 V linha R2", "marrom rigido encapado", "0,50 mm2", "Ramal do T2 e 24 V de servicos", "painel_interativo/src/data/maquete.js"),
    ("24 V linha R3", "cinza rigido encapado", "0,50 mm2", "Ramal do T3 e auxiliares", "painel_interativo/src/data/maquete.js"),
    ("0 V linha", "azul claro rigido encapado", "1,50 mm2", "Retorno comum dos tres ramais, 6,9 A", "painel_interativo/src/data/maquete.js"),
    ("12 V auxiliar", "amarelo / preto", "0,75 mm2", "T3/BD-AUX -> ventoinhas", "painel_interativo/src/data/fiacao.js"),
    ("5 V logica", "violeta ou vermelho / preto", "0,50 mm2", "T2/BD-5V -> Arduino, IHM e modulos", "painel_interativo/src/data/fiacao.js"),
    ("Sinais", "cinza, azul, amarelo ou verde", "0,25 mm2", "Sensores, RPM, I2C, UART e comandos", "painel_interativo/src/data/fiacao_etapa4.js"),
    ("I2C / 1-Wire", "par trançado blindado", "2 x 0,25 mm2", "Painel <-> camara; blindagem aterrada em um ponto", "painel_interativo/src/data/fiacao_etapa4.js"),
]

DIVERGENCIAS = [
    ("Critica", "Dimensao do painel", "BOM: 400 x 500 x 200 mm", "React: 500 x 500 x 200 mm", "Comprar/fabricar com 500 mm de largura; o layout atual precisa disso."),
    ("Critica", "PI-2 e ensaio", "BOM descreve itens espalhados e ainda cita 4 em alguns trechos", "React: 2 DUTs, CD74HC4067 e 1 INA219", "Comprar para 2 posicoes; manter 1 mux e 1 INA219."),
    ("Critica", "Ventoinhas internas", "BOM separa 2 fans de 60 mm e 2 fans de 40 mm", "React: 5 internas no mesmo grupo: 4 de circulacao + 1 do PTC", "Conferir fisicamente o tamanho das 5; a fiação atual trata as cinco em paralelo."),
    ("Alta", "BD-5V e BD-24V", "Trechos antigos do BOM/React antigo mostram 6 e 4 saidas", "React atual: BD-5V com 8 e BD-24V com 6", "Comprar os blocos maiores da aba Inventario React."),
    ("Alta", "IHM", "BOM antiga ainda cita Nextion e modulo SD separado em partes do texto", "React: ES3C28P com microSD integrado", "Nao comprar Nextion nem modulo SD separado; comprar conversor de nivel e rabicho Type-C."),
    ("Media", "Cores de fiação", "BOM resume positivos/negativos por cor", "React usa cor funcional: 24 V permanente laranja, 5 V violeta, sinal cinza/verde", "Seguir a aba Fiacao React e identificar cada cabo com anilha."),
    ("Media", "Geometria da camara", "BOM/documentos têm dimensoes revisadas diferentes", "React ainda conserva constantes antigas em camara.js, enquanto montagem/ligacoes refletem a revisao", "Nao cortar acrilico sem confirmar a lista de corte e o desenho final."),
]


def adicionar_aba_tabela(wb, nome, cabecalho, linhas, larguras=None):
    ws = wb.create_sheet(nome)
    for coluna, titulo in enumerate(cabecalho, start=1):
        c = ws.cell(row=1, column=coluna, value=titulo)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA
        if larguras:
            ws.column_dimensions[get_column_letter(coluna)].width = larguras[coluna - 1]
    for linha, valores in enumerate(linhas, start=2):
        for coluna, valor in enumerate(valores, start=1):
            c = ws.cell(row=linha, column=coluna, value=valor)
            c.border = BORDA
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if linha % 2 == 0:
                c.fill = PatternFill("solid", fgColor=CINZA_CLARO)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalho))}{max(1, len(linhas) + 1)}"
    return ws


def aplicar_react_como_fonte(itens: list[dict]) -> list[dict]:
    """Corrige a lista principal para seguir o inventario efetivo do React."""
    corrigidos = []
    for reg in itens:
        item = reg["item"].lower().translate(ACENTOS)
        if "fan 60" in item or "fan 40" in item:
            continue
        if "caixa de comando ou mdf" in item:
            reg["espec"] = "500 x 500 x 200 mm — dimensao exigida pelo layout React"
        elif "arduino mega 2560" in item:
            reg["qtd"] = "1"
            reg["espec"] = "1 controlador em uso; reserva opcional"
        elif "km1" in item:
            reg["qtd"] = "2"
            reg["espec"] = "1 conjunto em uso + 1 reserva; 24 Vcc, 2 contatos reversiveis de 10 A em CC"
        elif "modulo rele 1 canal 5 v" in item:
            reg["qtd"] = "3"
            reg["espec"] = "KA1 (NO), KA2 (NC) e KA3 (NO), optoacoplados, caixa DIN 6M; reservas opcionais"
        elif item.startswith("sensor ds18b20"):
            reg["espec"] = "1-Wire, 3 fios, colado no dissipador quente; nao fica no centro da camara"
        elif "placa ilhada (padrao" in item:
            reg["item"] = "Placa ilhada 9 x 15 cm para PI-1 + PI-2"
            reg["espec"] = "1 placa de 34 x 58 furos, cortada ao meio para formar as duas placas"
        elif "resistor 47" in item:
            reg["qtd"] = "2"
            reg["espec"] = "47 ohm, 1%, 1/4 W; R1 e R2 ativos na PI-2"
        elif "fusivel mini automotivo 500" in item:
            reg["qtd"] = "2"
            reg["espec"] = "100 mA por DUT; F-P1 e F-P2"
        elif "resistor 1,2" in item or "resistor 2,2" in item:
            reg["qtd"] = "1"
        elif "led 5 mm" in item:
            reg["qtd"] = "2"
        elif "micro-chave ou jumper" in item:
            reg["qtd"] = "2"
        elif "diodo 1n4007" in item:
            reg["qtd"] = "2"
        elif "ci uln2803a" in item:
            reg["qtd"] = "1"
        elif "cabo flexivel 1,5 mm² preto" in item:
            reg["item"] = "Cabo flexivel 1,5 mm² azul escuro"
            reg["espec"] = "750 V; 0 V comum conforme a codificacao do React"
        corrigidos.append(reg)

    corrigidos.append({
        "secao": "L.3 — Atuadores térmicos e ventilação",
        "item": "Ventoinha interna 12 V — grupo React",
        "qtd": "5",
        "espec": "Cinco unidades compatíveis com a geometria atual: 2 frias, 2 dos dutos e 1 do PTC; ligadas em paralelo",
        "obs": "Confirmar dimensao fisica antes da compra",
        "link_manual": "", "busca_ml": "", "busca_amz": "", "termo": "ventoinha 12v 40mm",
        "termo_en": "", "busca_ali": "",
    })
    return corrigidos


def gerar(itens: list[dict], destino: Path) -> None:
    wb = Workbook()

    # ───────────────────────── aba 1: lista de compras ─────────────────────
    ws = wb.active
    ws.title = "Lista de Compras"

    ws.merge_cells("A1:K1")
    ws["A1"] = "PROJETO INTEGRADOR — Planta Industrial Didática com Câmara Frigorífica"
    ws["A1"].font = Font(size=15, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = (
        f"Lista de materiais · gerada de 03_lista_materiais.md em "
        f"{datetime.now():%d/%m/%Y %H:%M} · arquitetura 127 V CA → 24 V CC → 12 / 5 / 3,3 V"
    )
    ws["A2"].font = Font(size=10, italic=True, color="44546A")
    ws["A2"].alignment = Alignment(horizontal="center")

    colunas = [
        ("#", 5),
        ("Item", 44),
        ("Qtd", 7),
        ("Especificação", 56),
        ("Observação", 32),
        ("🔎 Mercado Livre", 24),
        ("🔎 Amazon BR", 14),
        ("🌏 AliExpress (importar)", 26),
        ("Link escolhido (cole aqui)", 28),
        ("Preço unit. (R$)", 15),
        ("Total (R$)", 13),
    ]
    linha_cab = 4
    for indice, (titulo, largura) in enumerate(colunas, start=1):
        celula = ws.cell(row=linha_cab, column=indice, value=titulo)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = BORDA
        ws.column_dimensions[get_column_letter(indice)].width = largura
    ws.row_dimensions[linha_cab].height = 30
    ws.freeze_panes = f"A{linha_cab + 1}"

    linha = linha_cab + 1
    numero = 0
    secao_atual = None

    for reg in itens:
        if reg["secao"] != secao_atual:
            secao_atual = reg["secao"]
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=11)
            celula = ws.cell(row=linha, column=1, value=secao_atual)
            celula.font = Font(bold=True, size=11, color=AZUL_ESCURO)
            celula.fill = PatternFill("solid", fgColor=AMBAR)
            celula.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            celula.border = BORDA
            ws.row_dimensions[linha].height = 22
            linha += 1

        numero += 1
        valores = [numero, reg["item"], reg["qtd"], reg["espec"], reg["obs"],
                   "", "", "", reg["link_manual"], "", ""]
        for indice, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha, column=indice, value=valor)
            celula.border = BORDA
            celula.alignment = Alignment(
                horizontal="center" if indice in (1, 3, 10, 11) else "left",
                vertical="top",
                wrap_text=indice in (2, 4, 5),
            )
            if linha % 2 == 0:
                celula.fill = PatternFill("solid", fgColor=CINZA_CLARO)

        # Links de busca clicáveis. Sempre válidos: a busca mostra o que
        # existe hoje, enquanto um link de anúncio sai do ar.
        if reg["busca_ml"]:
            c = ws.cell(row=linha, column=6)
            c.value = f'=HYPERLINK("{reg["busca_ml"]}","🔎 {reg["termo"][:28]}")'
            c.font = Font(color="0563C1", underline="single", size=9)
        if reg["busca_amz"]:
            c = ws.cell(row=linha, column=7)
            c.value = f'=HYPERLINK("{reg["busca_amz"]}","🔎 Amazon")'
            c.font = Font(color="0563C1", underline="single", size=9)
        if reg["busca_ali"]:
            c = ws.cell(row=linha, column=8)
            c.value = f'=HYPERLINK("{reg["busca_ali"]}","🌏 {reg["termo_en"][:26]}")'
            c.font = Font(color="C55A11", underline="single", size=9)

        ws.cell(row=linha, column=11).value = (
            f"=IF(C{linha}*J{linha}=0,\"\",C{linha}*J{linha})"
        )
        ws.cell(row=linha, column=10).number_format = "#,##0.00"
        ws.cell(row=linha, column=11).number_format = "#,##0.00"
        linha += 1

    ws.cell(row=linha + 1, column=9, value="TOTAL GERAL").font = Font(bold=True, size=12)
    total = ws.cell(row=linha + 1, column=11, value=f"=SUM(K{linha_cab + 1}:K{linha - 1})")
    total.font = Font(bold=True, size=12, color="FFFFFF")
    total.fill = PatternFill("solid", fgColor=VERDE)
    total.number_format = "#,##0.00"
    total.border = BORDA

    ws.auto_filter.ref = f"A{linha_cab}:K{linha - 1}"

    adicionar_aba_tabela(
        wb,
        "Inventario React",
        ["Grupo", "Componente", "Qtd", "Especificacao conferida", "Fonte React"],
        REACT_INVENTARIO,
        [18, 34, 8, 64, 48],
    )
    adicionar_aba_tabela(
        wb,
        "Fiacao React",
        ["Circuito", "Cor(es)", "Bitola", "Uso", "Fonte React"],
        REACT_FIACAO,
        [22, 32, 16, 62, 48],
    )
    adicionar_aba_tabela(
        wb,
        "Divergencias",
        ["Prioridade", "Tema", "BOM/Documentacao", "React atual", "Decisao para compra"],
        DIVERGENCIAS,
        [14, 24, 52, 56, 64],
    )

    # ───────────────────────── aba 2: ordem de construção ──────────────────
    ws2 = wb.create_sheet("Ordem de Construção")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "ORDEM DE CONSTRUÇÃO — 5 camadas"
    ws2["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    etapas = [
        ("CAMADA 0", "Fundamentos", "Visão geral · Arquitetura de energia · BOM", "Entregável: BOM fechada e cálculos feitos"),
        ("CAMADA 1", "Maquete", "Base e chão de fábrica · Subestação e postes · Câmara térmica", "Entregável: cenário pronto, sem eletrônica"),
        ("CAMADA 2", "Painel", "Dimensionamento · Layout · Furação · Fixação nos trilhos", "Entregável: painel montado, sem um fio ligado"),
        ("CAMADA 3", "Elétrica", "Força e distribuição · Comando e proteções · Sinais e sensores", "Entregável: energizado e medido, sem firmware"),
        ("CAMADA 4", "Programação", "Firmware Arduino · ESP32, IHM e IoT", "Entregável: software gravado e testado em bancada"),
        ("CAMADA 5", "Integração", "Montagem final · Comissionamento · Ensaios", "Entregável: projeto funcionando e documentado"),
    ]
    for indice, titulo in enumerate(["Camada", "Nome", "Documentos", "Critério de aceitação"], start=1):
        celula = ws2.cell(row=3, column=indice, value=titulo)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
        celula.border = BORDA
    for largura, letra in zip((12, 18, 58, 52), "ABCD"):
        ws2.column_dimensions[letra].width = largura

    for i, etapa in enumerate(etapas, start=4):
        for j, valor in enumerate(etapa, start=1):
            celula = ws2.cell(row=i, column=j, value=valor)
            celula.border = BORDA
            celula.alignment = Alignment(vertical="top", wrap_text=True)
            if j == 1:
                celula.font = Font(bold=True, color=AZUL_ESCURO)
                celula.fill = PatternFill("solid", fgColor=AMBAR)

    # ───────────────────────── aba 3: dados do projeto ─────────────────────
    ws3 = wb.create_sheet("Dados do Projeto")
    ws3.merge_cells("A1:B1")
    ws3["A1"] = "NÚMEROS DO PROJETO"
    ws3["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28
    ws3.column_dimensions["A"].width = 46
    ws3.column_dimensions["B"].width = 46

    dados = [
        ("Arquitetura de energia", "Potência em 24 V — carga térmica sem conversor"),
        ("Tensão de entrada", "127 V CA"),
        ("Barramento de distribuição", "24 V CC (SELV)"),
        ("Potência instalada (fonte)", "240 W · 10 A"),
        ("Consumo calculado", "≈ 166 W · 6,9 A @ 24 V"),
        ("Folga da fonte", "1,44× (44 %) — a fonte de 150 W NÃO atende"),
        ("Corrente CA de entrada", "2,37 A"),
        ("Queda de tensão na linha dos postes", "0,21 V (0,86 %)"),
        ("Conversores", "2 × LM2596 com display (T2 e T3) · P1 é derivação direta"),
        ("Tensões derivadas", "24,0 V potência (direto) · 12,0 V auxiliar · 5,10 V comando · 3,3 V ESP32"),
        ("Fusíveis dos ramais", "F1 = 10 A · F2 = 2 A · F3 = 2 A (sem F4/F5, sem crowbar)"),
        ("Níveis de proteção", "5 (do disjuntor ao intertravamento por software)"),
        ("Volume útil da câmara", "5,0 litros (200 × 100 × 250 mm)"),
        ("Carga térmica calculada", "≈ 9,5 W"),
        ("Refrigeração", "2 × TEC1-12706 EM SÉRIE · 24 V · 6,0 A · 144 W"),
        ("Capacidade das Peltier a ΔT = 20 K", "≈ 60 W (margem de 6,3×)"),
        ("Dissipação no lado quente", "≈ 200 W — 2 conjuntos dissipador + cooler 80 mm"),
        ("Aquecimento", "PTC cerâmico de 24 V · 80 W · 3,3 A"),
        ("Isolamento", "XPS 30 mm + barreira de vapor · U = 0,86 W/m²·K"),
        ("Porta", "Dupla · U = 2,42 W/m²·K (não condensa)"),
        ("Base da maquete", "1500 × 500 mm · escala cenográfica 1:50"),
        ("Painel", "400 × 500 × 200 mm · 3 trilhos DIN"),
        ("Componentes discretos", "6 + 1 CI ULN2803 na PI-1 · 2 nos BTS7960 · 4 nos postes da maquete"),
        ("Sinalização", "4 sinaleiros 22 mm de 24 V (via ULN2803) · 4 LEDs brancos de 5 V na maquete"),
    ]
    for i, (rotulo, valor) in enumerate(dados, start=3):
        c1 = ws3.cell(row=i, column=1, value=rotulo)
        c2 = ws3.cell(row=i, column=2, value=valor)
        c1.font = Font(bold=True)
        for c in (c1, c2):
            c.border = BORDA
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=CINZA_CLARO)

    wb.save(destino)
    print(f"  Planilha gerada: {destino}")
    print(f"  {numero} itens em {len({i['secao'] for i in itens})} seções")


if __name__ == "__main__":
    gerar(aplicar_react_como_fonte(ler_bom(BOM_MD)), SAIDA)
