#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera arquivo Excel com lista de materiais COMPLETA do projeto.
Extrai dados do projeto_camara_v2.md e organiza por categoria.
Requer: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# BOM Completo com TODAS as quantidades do projeto_camara_v2.md
bom_completo = {
    "LÓGICA, INTERFACE E REDE": [
        ("Arduino Mega 2560 R3", "1", "ATmega2560, 54 GPIO, 4 UARTs nativas"),
        ("Shield expansão Mega", "1", "Sensor Shield V2.0 com bornes parafuso"),
        ("Tela Nextion Basic 3.2\"", "1", "NX4024T032 (400×240, TTL 5V)"),
        ("ESP32-WROOM-32U", "1", "30 pinos, antena IPEX externa"),
        ("DNLCB30 (base DIN ESP32)", "1", "Conversão 3.3V↔5V automática, 7–35V in"),
        ("Antena 2.4GHz IPEX→SMA", "1", "Pigtail com antena 3 dBi para painel"),
    ],
    
    "POTÊNCIA E ATUADORES TÉRMICOS (12V)": [
        ("Fonte ATX 500W", "1", "12V ≥20A, reutilizar fonte de PC"),
        ("Pastilha Peltier TEC1-12706", "1", "60W, ~6A @12V"),
        ("Aquecedor PTC cerâmico", "1", "12V, 50–100W, com aletas"),
        ("Fan 60×60 mm (12V)", "2", "Ventoinhas extras (PTC/Peltier)"),
        ("Fan 40×40 mm (12V)", "2", "Ventoinhas internas kit Peltier"),
    ],
    
    "DRIVERS E ACIONAMENTO": [
        ("BTS7960 Driver (módulo)", "2", "Dupla ponte-H, 43A contínuos, pino IS"),
        ("Suporte SPCI4 trilho DIN", "2", "ABS, fixa PCI 100×79mm, DIN 35mm"),
        ("Botão de emergência cogumelo", "1", "NF, 22mm, com trava, 2 contatos empilháveis"),
        ("Botão START", "1", "Verde, momentâneo NA, 22mm"),
        ("Botão STOP", "1", "Preto, momentâneo NF, 22mm"),
        ("Chave rotativa 0-1", "1", "22mm, força geral AC, lateral do painel"),
        ("Módulo Micro SD", "1", "SPI, 5V, compatível Arduino"),
        ("Módulo RTC DS3231", "1", "I²C, bateria CR2032, precisão ±2ppm"),
        ("Bateria CR2032", "1", "Para backup RTC"),
    ],
    
    "SENSORIAMENTO": [
        ("Sensor DS18B20", "1", "Digital 1-Wire, ±0.5°C, waterproof, centro câmara"),
        ("Sensor AM2315C", "1", "I²C, hum./temp, ±0.3°C / ±2% UR, carcaça fechada"),
        ("Resistor pull-up 4.7kΩ", "1", "1/4W para barramento 1-Wire"),
    ],
    
    "PAINEL MECÂNICO E SINALIZAÇÃO": [
        ("Trilho DIN 35mm", "1m", "Aço galvanizado"),
        ("Bornes parafuso DIN", "10", "2.5mm², ~32A cada"),
        ("LEDs sinalizadores 22mm", "4", "Verde/Azul/Amarelo/Vermelho, 12V"),
        ("Resistor 220Ω", "4", "Para LEDs série"),
        ("Capacitor 100nF", "2", "Filtro pino IS dos BTS7960"),
    ],
    
    "CABOS E CONECTORES": [
        ("Cabo flexível 2.5mm² (vm+pt)", "~5m", "Potência 12V"),
        ("Cabo flexível 0.5mm²", "~5m", "Lógica e alimentação periféricos"),
        ("Cabo 0.25–0.5mm²", "~10m", "Sinais, botões, sensores"),
        ("Prensa-cabo PG9", "4", "Entrada de cabos no painel"),
        ("Porta-fusível DIN", "1", "Para fusível 10A (ramal BTS)"),
        ("Fusível mini automotivo 10A", "1", "Ramal de potência BTS"),
    ],
    
    "ACRÍLICO E ESTRUTURA": [
        ("Acrílico transparente 5mm", "1", "Paredes câmara (6 peças cortadas a laser)"),
        ("Acrílico transparente 10mm", "1", "Porta frontal"),
        ("Acrílico transparente 3mm", "1", "Dutos externos (10 peças)"),
        ("Acrílico preto/cinza 3mm", "1", "Coberturas externas XPS (5 peças)"),
        ("Cola S-320 Sinteglas", "1", "250ml ou 1L, para colagem acrílico"),
        ("Silicone neutro transparente", "1", "1 bisnaga, vedação"),
        ("Perfil EPDM 5mm autoadesivo", "1m", "Vedação porta/bordas"),
        ("Dobradiça piano alumínio", "1", "75cm (cortar para 25cm)"),
        ("Fecho pressão caixa", "2", "Inox"),
        ("MDF 18mm", "1", "Backplate painel (300×450mm) + base (65×30cm)"),
        ("XPS 20mm", "1", "Isolamento câmara (~50×50cm)"),
        ("Sílica gel indicadora", "2 sachês", "50g, controle umidade transporte"),
    ],
    
    "PARAFUSOS E FIXAÇÕES": [
        ("Parafuso M3×8mm inox", "Kit", "Para dobradiça/diversos"),
        ("Parafuso M4×15mm", "Kit", "Para câmara/diversos"),
        ("Parafuso M5×20mm", "Kit", "Para trilho DIN/estrutura"),
        ("Pé borracha autoadesivo", "4", "Para base da maquete"),
    ],
}

def gerar_excel_completo(nome_arquivo="BOM_Completo_Camara_Frigorifica.xlsx"):
    """Gera Excel com BOM completo por categorias."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Completo"
    
    # Estilos
    font_titulo = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    fill_titulo = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    
    font_categoria = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    fill_categoria = PatternFill(start_color='0056B3', end_color='0056B3', fill_type='solid')
    
    font_cabecalho = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_cabecalho = PatternFill(start_color='0056B3', end_color='0056B3', fill_type='solid')
    
    fill_linhas_pares = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Título geral
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = 'LISTA COMPLETA DE MATERIAIS (BOM) — MINI CÂMARA FRIGORÍFICA'
    cell.font = font_titulo
    cell.fill = fill_titulo
    cell.alignment = alignment_center
    ws.row_dimensions[1].height = 25
    
    # Subtítulo
    ws.merge_cells('A2:E2')
    cell = ws['A2']
    cell.value = f"Projeto Integrador Técnico em Eletrotécnica | Versão 3.0"
    cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
    cell.alignment = alignment_center
    
    ws.merge_cells('A3:E3')
    cell = ws['A3']
    cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
    cell.alignment = alignment_center
    
    linha_atual = 5
    
    # Cabeçalho da tabela
    cabecalhos = ['Item', 'Qtd', 'Observação / Especificação', 'Link de Compra', 'Preço Unitário']
    for col_num, cabecalho in enumerate(cabecalhos, 1):
        cell = ws.cell(row=linha_atual, column=col_num)
        cell.value = cabecalho
        cell.font = font_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alignment_center
        cell.border = border_thin
    
    ws.row_dimensions[linha_atual].height = 18
    linha_atual += 1
    
    # Adicionar dados por categoria
    total_itens = 0
    for categoria, itens in bom_completo.items():
        # Linha de categoria
        ws.merge_cells(f'A{linha_atual}:E{linha_atual}')
        cell = ws[f'A{linha_atual}']
        cell.value = categoria
        cell.font = font_categoria
        cell.fill = fill_categoria
        cell.alignment = alignment_left
        ws.row_dimensions[linha_atual].height = 18
        linha_atual += 1
        
        # Itens da categoria
        for idx, (item, qtd, obs) in enumerate(itens):
            total_itens += 1
            
            # Item
            cell = ws.cell(row=linha_atual, column=1)
            cell.value = item
            cell.alignment = alignment_left
            cell.border = border_thin
            
            # Quantidade
            cell = ws.cell(row=linha_atual, column=2)
            cell.value = qtd
            cell.alignment = alignment_center
            cell.border = border_thin
            
            # Observação
            cell = ws.cell(row=linha_atual, column=3)
            cell.value = obs
            cell.alignment = alignment_left
            cell.border = border_thin
            
            # Link (vazio para preenchimento)
            cell = ws.cell(row=linha_atual, column=4)
            cell.value = ""
            cell.alignment = alignment_left
            cell.border = border_thin
            
            # Preço (vazio para preenchimento)
            cell = ws.cell(row=linha_atual, column=5)
            cell.value = ""
            cell.alignment = alignment_center
            cell.border = border_thin
            
            # Cor alternada dentro de cada categoria
            if idx % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=linha_atual, column=col).fill = fill_linhas_pares
            
            linha_atual += 1
        
        # Espaço entre categorias
        linha_atual += 1
    
    # Rodapé com total
    linha_atual += 1
    ws.merge_cells(f'A{linha_atual}:E{linha_atual}')
    cell = ws[f'A{linha_atual}']
    cell.value = f"TOTAL DE ITENS: {total_itens}"
    cell.font = Font(name='Calibri', size=11, bold=True, color='003366')
    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    cell.alignment = alignment_center
    cell.border = border_thin
    
    # Ajustar largura de colunas
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 14
    
    # Congelar cabeçalho
    ws.freeze_panes = 'A6'
    
    # Salvar workbook
    wb.save(nome_arquivo)
    print(f"✓ Arquivo Excel COMPLETO gerado: {nome_arquivo}")
    print(f"  Total de itens: {total_itens}")
    print(f"  Categorias: {len(bom_completo)}")
    print(f"  Colunas prontas para preenchimento: Link de Compra e Preço Unitário")

if __name__ == "__main__":
    try:
        gerar_excel_completo()
    except ImportError:
        print("Erro: openpyxl não está instalado.")
        print("Execute: pip install openpyxl")
    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
