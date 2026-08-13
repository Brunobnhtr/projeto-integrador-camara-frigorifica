#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera arquivo Excel com lista de materiais (BOM) do projeto de câmara frigorífica.
Requer: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Dados de materiais
materiais = [
    ("Arduino Mega 2560 R3", "1", "CLP central — ATmega2560"),
    ("Shield expansão Mega (bornes)", "1", "Sensor shield / bornes parafuso"),
    ("Tela Nextion Basic 3.2\" (NX4024T032)", "1", "IHM TTL 5V"),
    ("ESP32-WROOM-32U", "1", "Módulo Wi‑Fi (antena IPEX opcional)"),
    ("DNLCB30 (base DIN para ESP32)", "1", "Base com conversão 7–35V → 3.3/5V"),
    ("Antena 2.4GHz IPEX→SMA", "1", "Pigtail p/ painel (3 dBi)"),
    ("Fonte ATX 500W (12V ≥20A)", "1", "Alimentação principal (usar ATX)"),
    ("Chave rotativa 0-1 (22mm)", "1", "Único comando AC — liga/desliga a fonte (≥6A AC). Sem disjuntor/seccionadora"),
    ("Pastilha Peltier (TEC1-12706)", "1", "~60W, ~6A @12V"),
    ("Aquecedor PTC cerâmico 12V", "1", "50–100W, com aletas"),
    ("Fan 60×60 mm (12V)", "2", "Ventoinhas extras (PTC/Peltier)"),
    ("Fan 40×40 mm (12V)", "2", "Ventoinhas internas kit Peltier"),
    ("BTS7960 Driver (módulo)", "2", "Ponte H para 12V/43A, pino IS para diagnóstico"),
    ("Suporte SPCI4 trilho DIN", "2", "Fixação módulos no trilho DIN"),
    ("Botão de emergência (cogumelo 22mm)", "1", "NF com trava (2 contatos empilháveis)"),
    ("Botões START / STOP 22mm", "2", "START NA / STOP NF"),
    ("Módulo Micro SD (SPI)", "1", "Logging local"),
    ("Módulo RTC DS3231", "1", "I²C, bateria CR2032"),
    ("Bateria CR2032", "1", "Backup RTC"),
    ("Sensor DS18B20 (waterproof)", "1", "1‑Wire, centro da câmara"),
    ("Sensor AM2315C (hum./temp)", "1", "I²C (opcional)"),
    ("Resistor 4.7kΩ (pull-up)", "1", "1‑Wire pull‑up"),
    ("Trilho DIN 35 mm", "1", "1 m, cortar conforme necessidade"),
    ("Bornes parafuso DIN (2.5 mm²)", "10", "Distribuição 12V/5V/GND"),
    ("LEDs sinalizadores 22 mm (V, B, Am, R)", "4", "12V, cores variadas"),
    ("Cabos flexíveis 2.5 mm² (preto/vermelho)", "Vários", "Potência 12V"),
    ("Cabos flexíveis 0.5 mm²", "Vários", "Lógica e alimentação periféricos"),
    ("Cabos 0.25–0.5 mm²", "Vários", "Sinais, botões e sensores"),
    ("Porta‑fusível para trilho DIN (fusível 10A)", "1", "Para ramal BTS"),
    ("Fusível mini automotivo 10A", "1", "Ramal de potência BTS"),
    ("Prensa-cabo PG9", "4", "Entrada de cabos no painel"),
    ("MDF 18 mm (painel / base)", "1", "Backplate e base da maquete"),
    ("Acrílico transpar. 5 mm", "Várias", "Paredes da câmara (ver lista acrílico)"),
    ("Acrílico transpar. 10 mm", "1", "Porta frontal"),
    ("Acrílico transpar. 3 mm", "Várias", "Dutos externos"),
    ("Acrílico preto 3 mm", "Várias", "Coberturas externas"),
    ("Cola S-320 / Silicone neutro", "1", "Fixação acrílico e vedação"),
    ("Perfil EPDM 5 mm autoadesivo", "1 m", "Vedação porta / bordas"),
    ("Dobradiça piano alumínio (cortar)", "1", "Para porta frontal"),
    ("Parafusos M3, M4, M5 kits", "Vários", "Fixações gerais"),
    ("Sílica gel indicadora", "2 sachês", "Controle umidade transporte"),
]

def gerar_excel_materiais(nome_arquivo="BOM_Camara_Frigorifica.xlsx"):
    """Gera arquivo Excel com tabela de materiais."""
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Materiais"
    
    # Estilos
    font_titulo = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    fill_titulo = PatternFill(start_color='0056B3', end_color='0056B3', fill_type='solid')
    
    font_cabecalho = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_cabecalho = PatternFill(start_color='0056B3', end_color='0056B3', fill_type='solid')
    
    fill_linhas_pares = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Título geral
    ws.merge_cells('A1:D1')
    cell_titulo = ws['A1']
    cell_titulo.value = 'LISTA DE MATERIAIS (BOM) — CÂMARA FRIGORÍFICA'
    cell_titulo.font = font_titulo
    cell_titulo.fill = fill_titulo
    cell_titulo.alignment = alignment_center
    ws.row_dimensions[1].height = 25
    
    # Data de geração
    ws.merge_cells('A2:D2')
    cell_data = ws['A2']
    cell_data.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    cell_data.font = Font(name='Calibri', size=9, italic=True, color='666666')
    cell_data.alignment = alignment_center
    
    # Cabeçalho da tabela (linha 4)
    cabecalhos = ['Item', 'Qtd', 'Observação', 'Link de Compra']
    for col_num, cabecalho in enumerate(cabecalhos, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = cabecalho
        cell.font = font_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alignment_center
        cell.border = border_thin
    
    ws.row_dimensions[4].height = 20
    
    # Adicionar dados
    for row_num, (item, qtd, obs) in enumerate(materiais, start=5):
        # Item
        cell = ws.cell(row=row_num, column=1)
        cell.value = item
        cell.alignment = alignment_left
        cell.border = border_thin
        
        # Quantidade
        cell = ws.cell(row=row_num, column=2)
        cell.value = qtd
        cell.alignment = alignment_center
        cell.border = border_thin
        
        # Observação
        cell = ws.cell(row=row_num, column=3)
        cell.value = obs
        cell.alignment = alignment_left
        cell.border = border_thin
        
        # Link de Compra (vazio para preenchimento)
        cell = ws.cell(row=row_num, column=4)
        cell.value = ""
        cell.alignment = alignment_left
        cell.border = border_thin
        
        # Aplicar cor alternada
        if row_num % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row_num, column=col).fill = fill_linhas_pares
    
    # Ajustar largura de colunas
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    
    # Congelar cabeçalho
    ws.freeze_panes = 'A5'
    
    # Salvar workbook
    wb.save(nome_arquivo)
    print(f"✓ Arquivo Excel gerado: {nome_arquivo}")
    print(f"  Total de itens: {len(materiais)}")
    print(f"  Coluna D está pronta para você inserir os links de compra")

if __name__ == "__main__":
    try:
        gerar_excel_materiais()
    except ImportError:
        print("Erro: openpyxl não está instalado.")
        print("Execute: pip install openpyxl")
    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
